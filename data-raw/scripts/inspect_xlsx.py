"""Quick inspection of all xlsx files in a directory.

Lists sheets and prints the first ~20 rows of each sheet as raw cell
values, plus the bounds of the sheet. Used to drive the design of the
ETL build scripts without needing R installed.
"""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook


def preview_sheet(ws, n_rows: int = 20, n_cols: int = 15) -> None:
    rows = list(ws.iter_rows(values_only=True))
    print(f"    rows={ws.max_row} cols={ws.max_column}  (showing first {min(n_rows, len(rows))} x {n_cols})")
    for r in rows[:n_rows]:
        cells = ["" if v is None else str(v) for v in r[:n_cols]]
        # truncate long cells
        cells = [c[:30] + "…" if len(c) > 30 else c for c in cells]
        print("      " + " | ".join(cells))


def main(folder: str) -> None:
    folder = Path(folder)
    files = sorted(folder.glob("*.xlsx"))
    for f in files:
        print("=" * 80)
        print(f"FILE: {f.name}")
        print("=" * 80)
        try:
            wb = load_workbook(f, read_only=True, data_only=True)
        except Exception as e:
            print(f"  ERROR: {e}")
            continue
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            print(f"  SHEET: {sheet!r}")
            preview_sheet(ws)
            print()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("usage: python inspect_xlsx.py <folder>")
        sys.exit(1)
    main(sys.argv[1])
