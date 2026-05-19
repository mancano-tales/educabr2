"""Inspect a single sheet in full detail: all columns, multi-row headers, sample data."""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook


def inspect(path: str, sheet: str, n_rows: int = 5) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    print(f"SHEET {sheet!r}: rows={ws.max_row} cols={ws.max_column}")
    rows = list(ws.iter_rows(values_only=True))
    print(f"\n--- HEADER ROWS (first 3) ---")
    for i, r in enumerate(rows[:3]):
        for j, v in enumerate(r):
            if v is not None:
                print(f"  row{i+1} col{j+1}: {v!r}")
        print()
    print(f"--- DATA SAMPLE (rows 4-{4+n_rows-1}) ---")
    for r in rows[3:3+n_rows]:
        print("  " + " | ".join("" if v is None else f"{v}" for v in r))
    print(f"\n--- LAST ROW (row {ws.max_row}) ---")
    print("  " + " | ".join("" if v is None else f"{v}" for v in rows[-1]))


if __name__ == "__main__":
    inspect(sys.argv[1], sys.argv[2], int(sys.argv[3]) if len(sys.argv) > 3 else 5)
